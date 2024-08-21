# -*- coding: utf-8 -*-
"""
Created on Tue Sep  5 16:11:04 2023

@author: snazrul
This script looks at the refund table in the database and gathers all the refunds for the give time frame and saves them in an excel file,
cleans them and performs calculation to be sent to the investor
"""

import pandas as pd
import mysql.connector
# import win32com.client as win32
# from datetime import datetime
# import os
# from pathlib import Path
from datetime import date, timedelta

today = date.today()
print("Today's date:", today)
today.strftime("%Y.%m.%d")

days_until_friday = (today.weekday() - 4) % 7

# Subtract the difference to get the date of the last Friday
last_friday = today - timedelta(days=days_until_friday)

# Format the date as a string (e.g., '2023-09-08')
last_friday_str = last_friday.strftime('%m%d%Y')

print("Last Friday's date:", last_friday_str)



connection = mysql.connector.connect(host= 'db.company.com',
                                user= '',
                                password= '',
                                database='' 
                                )

def get_query(filename):
    '''Read the Query from the SQL file and change the reportDate. Returns the Query in string format'''
    with open(filename,'r') as q:
        query = q.read()
    print(filename)
    return query

sql_file = 'C:/Users/snazrul/Desktop/Refunds_Code/borrower_refund_backup_v2.sql'

df_all = pd.read_sql(get_query(sql_file), connection)

writer = pd.ExcelWriter("C:/Users/snazrul/Desktop/Refunds_Code/%s refunds.xlsx" %(today.strftime("%Y.%m.%d")), engine='xlsxwriter')


backup_columns = df_all[['company ID', 'GDS Reference ID','School','School Refund Amount Received','Refund Percentage','Refund Risk Share', 'Refund O_fee Manually'
                         ,'company write off amount - servicer non-cash','servicer Cashout - Without  O Fee', 'Total refund amount','Effective Date', 'Withdrawal Date','refund_incomplete_reason']
                        ].rename(columns={'School Refund Amount Received': 'Refund from school', 'Refund Percentage':'Refund %'})


sum_row = backup_columns[['Refund from school',
                          'Refund Risk Share', 'Refund O_fee Manually',
                          'company write off amount - servicer non-cash',
                          'servicer Cashout - Without  O Fee',
                          'Total refund amount']].sum()

sum_row_dict = sum_row.to_dict()

df_to_add = pd.DataFrame(sum_row_dict, index=[0])

# Concatenate the original DataFrame with the DataFrame created from the dictionary
result_df = pd.concat([backup_columns, df_to_add], ignore_index=True)

# result_df.to_csv("C:/Users/snazrul/Desktop/tt1.csv")


# append the sum row to the DataFrame
#backup_columns = backup_columns.append(sum_row, ignore_index=True)

# backup_columns = backup_columns.append(sum_row, ignore_index=True)
#backup_columns = pd.concat([backup_columns,sum_row],ignore_index=True)

cash_columns = df_all[['GDS Reference ID','School','SSN','Sequence #', 'Disbursement Date','servicer Cashout - Without  O Fee','Transaction','Instituition ID','Withdrawal Date']
                      ].rename(columns={'Withdrawal Date': 'Seperation Date', 'School Refund Amount Received':'Refund/Cancel Amount'})

cash_columns.insert(loc=cash_columns.columns.get_loc('Instituition ID') + 1, column='Advice', value=2)
cash_columns.insert(loc=cash_columns.columns.get_loc('Instituition ID') + 1, column='Transmittal Notice', value=None)



non_cash_columns = df_all[df_all['Advice']==3][['company ID','GDS Reference ID','School','SSN','Sequence #', 'Disbursement Date','company write off amount - servicer non-cash',
                                                'Transaction','Instituition ID','Withdrawal Date']].rename(columns={'Withdrawal Date': 'Seperation Date', 'company write off amount - servicer non-cash':'Refund/Cancel Amount'})

non_cash_columns['Transaction'] = 1027
non_cash_columns.insert(loc=non_cash_columns.columns.get_loc('Instituition ID') + 1, column='Advice', value=3)
non_cash_columns.insert(loc=non_cash_columns.columns.get_loc('Instituition ID') + 1, column='Transmittal Notice', value=None)


result_df.to_excel(writer, sheet_name='Backup', index=False)
cash_columns.to_excel(writer, sheet_name='Cash', startrow = 5, index=False)
non_cash_columns.to_excel(writer, sheet_name='Non_Cash', startrow = 5, index=False)



# Save the Excel file
# writer.save()

writer.close()
#company writeoff amount, anything that is not 0 is non cash


import openpyxl

# Open an existing Excel workbook or create a new one
workbook = openpyxl.load_workbook("C:/Users/snazrul/Desktop/Refunds_Code/%s refunds.xlsx" %(today.strftime("%Y.%m.%d")))  # Replace 'your_excel_file.xlsx' with your file name


workbook.active = workbook['Non_Cash']

nc_string_1 = "File Naming Convention: company.Refund.Roster.{}.csv".format(last_friday_str)
nc_string_2 = "File Format: CSV"
nc_string_3 = "Non Cash Cancellation - 1027"

sheet = workbook['Non_Cash']
sheet['A1'] = nc_string_1
sheet['A2'] = nc_string_2
sheet['A3'] = nc_string_3


sheet2 = workbook['Cash']

c_string_3 = "Cancellation (1045)/Refund (1040)"

sheet2['A1'] = nc_string_1
sheet2['A2'] = nc_string_2
sheet2['A3'] = c_string_3

# backup = workbook['Backup']
# row_n = len(backup_columns)+1
# row_n_string = "{}".format(row_n)
# cell_v = 'D'+row_n_string

# num_rows = backup.max_row

# # Sum the values in column D
# sum_column_d = sum(sheet.cell(row=i, column=4).value for i in range(1, num_rows + 1) if sheet.cell(row=i, column=4).value is not None)


# backup['{}'.format(cell_v)] = backup_columns[['Refund from school',]].sum().tolist()
# Save the changes to the Excel file
# workbook.save("C:/Refunds_Code/%s refunds.xlsx" %(today.strftime("%Y.%m.%d")))
workbook.save("C:/Users/snazrul/Desktop/Refunds_Code/%s refunds.xlsx" %(today.strftime("%Y.%m.%d")))
# Close the workbook
workbook.close()




















